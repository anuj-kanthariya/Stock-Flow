"""Products router."""
import math
from fastapi import APIRouter, Depends, Query, HTTPException, status, File, UploadFile
import os
import uuid
import aiofiles
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, or_
from sqlalchemy.orm import joinedload

from app.database.database import get_db
from app.schemas.schemas import (
    ProductCreate,
    ProductUpdate,
    ProductResponse,
    ProductListResponse,
    PaginationMeta,
)
from app.core.dependencies import get_current_active_user, get_current_owner
from app.models.models import User, Product, Category

router = APIRouter()


@router.get("/", response_model=ProductListResponse)
async def list_products(
    page: int = Query(1, ge=1),
    limit: int = Query(20, ge=1, le=100),
    search: str = Query(""),
    category_id: str = Query(""),
    sort_by: str = Query("name"),
    is_active: bool = Query(True),
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db),
):
    """List products with pagination and filtering."""
    query = (
        select(Product)
        .options(joinedload(Product.category))
        .where(Product.is_active == is_active)
        .where(Product.owner_id == current_user.id)
    )

    if search:
        query = query.where(
            or_(
                Product.name.ilike(f"%{search}%"),
                Product.brand.ilike(f"%{search}%"),
                Product.sku.ilike(f"%{search}%")
            )
        )

    if category_id:
        query = query.where(Product.category_id == category_id)

    if sort_by == "name":
        query = query.order_by(Product.name.asc())
    elif sort_by == "-name":
        query = query.order_by(Product.name.desc())
    elif sort_by == "stock":
        query = query.order_by(Product.stock_quantity.asc())
    elif sort_by == "-stock":
        query = query.order_by(Product.stock_quantity.desc())
    else:
        query = query.order_by(Product.created_at.desc())

    # Total count
    count_query = select(func.count()).select_from(query.subquery())
    total_res = await db.execute(count_query)
    total = total_res.scalar() or 0

    # Pagination
    offset = (page - 1) * limit
    query = query.offset(offset).limit(limit)

    result = await db.execute(query)
    products = result.scalars().all()

    # Map to response format
    data = []
    for prod in products:
        prod_data = prod.__dict__.copy()
        prod_data["category_name"] = prod.category.name if prod.category else ""
        data.append(prod_data)

    total_pages = math.ceil(total / limit) if limit else 1

    return ProductListResponse(
        meta=PaginationMeta(
            page=page,
            limit=limit,
            total=total,
            total_pages=total_pages,
        ),
        data=data,
    )


@router.post("/", response_model=ProductResponse, status_code=status.HTTP_201_CREATED)
async def create_product(
    data: ProductCreate,
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db),
):
    """Create a new product."""
    # Validate category
    cat_res = await db.execute(
        select(Category)
        .where(Category.id == data.category_id)
        .where(Category.owner_id == current_user.id)
    )
    category = cat_res.scalar_one_or_none()
    if not category:
        raise HTTPException(status_code=400, detail="Category does not exist.")

    # Validate SKU uniqueness
    if data.sku and data.sku.strip():
        sku_res = await db.execute(
            select(Product)
            .where(Product.sku == data.sku)
            .where(Product.owner_id == current_user.id)
        )
        if sku_res.scalar_one_or_none():
            raise HTTPException(status_code=400, detail="Product with this SKU already exists.")

    # Validate Barcode uniqueness
    if data.barcode and data.barcode.strip():
        barcode_res = await db.execute(
            select(Product)
            .where(Product.barcode == data.barcode)
            .where(Product.owner_id == current_user.id)
        )
        if barcode_res.scalar_one_or_none():
            raise HTTPException(status_code=400, detail="Product with this Barcode already exists.")

    new_product = Product(**data.model_dump(), owner_id=current_user.id)
    db.add(new_product)
    await db.commit()
    await db.refresh(new_product)

    # Attach category name for response
    prod_data = new_product.__dict__.copy()
    prod_data["category_name"] = category.name
    return prod_data


@router.get("/{product_id}", response_model=ProductResponse)
async def get_product(
    product_id: str,
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db),
):
    """Get a product by ID."""
    query = (
        select(Product)
        .options(joinedload(Product.category))
        .where(Product.id == product_id)
        .where(Product.owner_id == current_user.id)
    )
    result = await db.execute(query)
    product = result.scalar_one_or_none()

    if not product:
        raise HTTPException(status_code=404, detail="Product not found.")

    prod_data = product.__dict__.copy()
    prod_data["category_name"] = product.category.name if product.category else ""
    return prod_data


@router.patch("/{product_id}", response_model=ProductResponse)
async def update_product(
    product_id: str,
    data: ProductUpdate,
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db),
):
    """Update a product."""
    query = (
        select(Product)
        .options(joinedload(Product.category))
        .where(Product.id == product_id)
        .where(Product.owner_id == current_user.id)
    )
    result = await db.execute(query)
    product = result.scalar_one_or_none()

    if not product:
        raise HTTPException(status_code=404, detail="Product not found.")

    update_data = data.model_dump(exclude_unset=True)

    # Validate Category
    if "category_id" in update_data:
        cat_res = await db.execute(
            select(Category)
            .where(Category.id == update_data["category_id"])
            .where(Category.owner_id == current_user.id)
        )
        if not cat_res.scalar_one_or_none():
            raise HTTPException(status_code=400, detail="Category does not exist.")

    # Validate SKU uniqueness
    if "sku" in update_data and update_data["sku"] and update_data["sku"].strip() and update_data["sku"] != product.sku:
        sku_res = await db.execute(
            select(Product)
            .where(Product.sku == update_data["sku"])
            .where(Product.owner_id == current_user.id)
        )
        if sku_res.scalar_one_or_none():
            raise HTTPException(status_code=400, detail="Product with this SKU already exists.")

    # Validate Barcode uniqueness
    if "barcode" in update_data and update_data["barcode"] and update_data["barcode"].strip() and update_data["barcode"] != product.barcode:
        barcode_res = await db.execute(
            select(Product)
            .where(Product.barcode == update_data["barcode"])
            .where(Product.owner_id == current_user.id)
        )
        if barcode_res.scalar_one_or_none():
            raise HTTPException(status_code=400, detail="Product with this Barcode already exists.")

    for key, value in update_data.items():
        setattr(product, key, value)

    await db.commit()
    await db.refresh(product)

    # Re-fetch category if it changed
    cat_name = product.category.name if product.category else ""
    if "category_id" in update_data:
        cat_res = await db.execute(select(Category).where(Category.id == product.category_id))
        category = cat_res.scalar_one_or_none()
        cat_name = category.name if category else ""

    prod_data = product.__dict__.copy()
    prod_data["category_name"] = cat_name
    return prod_data


MAX_FILE_SIZE = 5 * 1024 * 1024  # 5 MB
ALLOWED_EXTENSIONS = {".jpg", ".jpeg", ".png", ".webp"}

@router.post("/{product_id}/image")
async def upload_product_image(
    product_id: str,
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db),
):
    """Upload and update a product image."""
    result = await db.execute(
        select(Product)
        .where(Product.id == product_id)
        .where(Product.owner_id == current_user.id)
    )
    product = result.scalar_one_or_none()
    
    if not product:
        raise HTTPException(status_code=404, detail="Product not found.")

    # Validate file extension
    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in ALLOWED_EXTENSIONS:
        raise HTTPException(status_code=400, detail="Invalid image type. Allowed: jpg, jpeg, png, webp")

    # Validate file size
    # We do not read the entire file here, the storage provider handles it,
    # but we can check the size via file.size if available, or just pass it along.
    # To keep it simple, we just pass the file to the storage provider.
    
    # Delete old image if it exists (local only currently, a full implementation would delete from remote)
    if product.image_url and product.image_url.startswith("/uploads/products/"):
        old_file_path = product.image_url.lstrip("/")
        if os.path.exists(old_file_path):
            try:
                os.remove(old_file_path)
            except Exception:
                pass

    # Save new image via storage abstraction
    from app.utils.storage import storage
    image_url = await storage.upload(file, "products", product_id)
    
    # Update product
    product.image_url = image_url
    await db.commit()
    
    return {"image_url": image_url}


@router.delete("/{product_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_product(
    product_id: str,
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db),
):
    """Soft delete a product."""
    result = await db.execute(
        select(Product)
        .where(Product.id == product_id)
        .where(Product.owner_id == current_user.id)
    )
    product = result.scalar_one_or_none()

    if not product:
        raise HTTPException(status_code=404, detail="Product not found.")

    product.is_active = False
    await db.commit()
    return None

import io
import openpyxl
from openpyxl.styles import Font, PatternFill
from fastapi import UploadFile, File, Form, HTTPException, status
from fastapi.responses import StreamingResponse
from app.schemas.schemas import (
    ProductImportRow,
    ProductImportPreviewResponse,
    ProductImportExecuteRequest,
    ProductImportExecuteResponse
)
from app.models.models import Category
from decimal import Decimal, InvalidOperation

@router.get("/import/template")
async def download_import_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Products"

    headers = [
        "Product Name", "SKU", "Category", 
        "Selling Price", "Cost Price", "Stock", "Unit"
    ]
    
    # Write headers with styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20

    # Sample row
    sample_data = ["Sample Product", "SKU-001", "Electronics", 999.99, 800.00, 50, "pcs"]
    for col, value in enumerate(sample_data, start=1):
        ws.cell(row=2, column=col, value=value)

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="StockFlow_Products_Template.xlsx"'}
    )

@router.post("/import/preview", response_model=ProductImportPreviewResponse)
async def preview_excel_import(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db)
):
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Invalid file format. Please upload an Excel file.")

    contents = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
        ws = wb.active
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Failed to read Excel file: {str(e)}")

    # Get existing SKUs
    existing_products_result = await db.execute(
        select(Product.sku).where(
            Product.owner_id == current_user.id,
            Product.sku.isnot(None)
        )
    )
    existing_skus = {sku.lower() for sku in existing_products_result.scalars().all()}
    
    # Get existing Categories
    existing_categories_result = await db.execute(
        select(Category.name).where(Category.owner_id == current_user.id)
    )
    existing_category_names = {name.lower() for name in existing_categories_result.scalars().all()}

    rows_data = []
    headers = [str(cell.value).strip().lower() if cell.value else "" for cell in ws[1]]
    
    def find_col(possible_names):
        for name in possible_names:
            if name in headers:
                return headers.index(name)
        return -1

    idx_name = find_col(["product name", "product"])
    idx_sku = find_col(["sku"])
    idx_cat = find_col(["category"])
    idx_sell = find_col(["selling price", "price"])
    idx_cost = find_col(["cost price", "cost"])
    idx_stock = find_col(["stock", "quantity"])
    idx_unit = find_col(["unit"])

    # Ensure minimum required columns exist
    if idx_name == -1 or idx_cat == -1 or idx_sell == -1:
         raise HTTPException(status_code=400, detail="Missing required columns (Product Name, Category, Price).")

    valid_count = 0
    invalid_count = 0
    duplicate_count = 0
    new_cats = set()

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # Skip completely empty rows
        if not any(row):
            continue
            
        errors = []
        name = str(row[idx_name]).strip() if idx_name >= 0 and row[idx_name] is not None else ""
        sku = str(row[idx_sku]).strip() if idx_sku >= 0 and row[idx_sku] is not None else ""
        category = str(row[idx_cat]).strip() if idx_cat >= 0 and row[idx_cat] is not None else ""
        selling_price_raw = row[idx_sell] if idx_sell >= 0 else None
        cost_price_raw = row[idx_cost] if idx_cost >= 0 else None
        stock_raw = row[idx_stock] if idx_stock >= 0 else 0
        unit = str(row[idx_unit]).strip() if idx_unit >= 0 and row[idx_unit] is not None else "pcs"

        if not name:
            errors.append("Product Name is missing")
        if not category:
            errors.append("Category is missing")

        import re
        def extract_number(val):
            if val is None: return None
            s = str(val).strip().replace(',', '')
            match = re.search(r'-?(?:\d+(?:\.\d*)?|\.\d+)', s)
            if match:
                return match.group(0)
            return None

        # Validate Selling Price
        selling_price = None
        sp_clean = extract_number(selling_price_raw)
        if sp_clean is not None:
            try:
                selling_price = Decimal(sp_clean)
                if selling_price <= 0:
                    errors.append("Selling price must be > 0")
            except (InvalidOperation, TypeError, ValueError):
                errors.append("Invalid Selling Price")
        else:
            errors.append("Invalid Selling Price")

        # Validate Cost Price
        cost_price = None
        if cost_price_raw is not None and str(cost_price_raw).strip() != "":
            cp_clean = extract_number(cost_price_raw)
            if cp_clean is not None:
                try:
                    cost_price = Decimal(cp_clean)
                    if cost_price < 0:
                        errors.append("Cost price cannot be negative")
                except (InvalidOperation, TypeError, ValueError):
                    errors.append("Invalid Cost Price")
            else:
                errors.append("Invalid Cost Price")

        # Validate Stock
        stock = 0
        st_clean = extract_number(stock_raw)
        if st_clean is not None:
            try:
                stock = int(float(st_clean))
                if stock < 0:
                    errors.append("Stock cannot be negative")
            except (TypeError, ValueError):
                errors.append("Invalid Stock value")
        else:
            errors.append("Invalid Stock value")

        status = "valid"
        if errors:
            status = "invalid"
            invalid_count += 1
        elif sku and sku.lower() in existing_skus:
            status = "duplicate"
            errors.append("SKU already exists")
            duplicate_count += 1
        else:
            valid_count += 1
            if sku:
                # Add to set so we catch duplicates within the same excel file
                existing_skus.add(sku.lower())

        if category and category.lower() not in existing_category_names:
            new_cats.add(category.lower())

        rows_data.append(ProductImportRow(
            row_number=row_idx,
            name=name,
            sku=sku if sku else None,
            category=category,
            selling_price=selling_price or Decimal("0"),
            cost_price=cost_price,
            stock=stock,
            unit=unit if unit else "pcs",
            status=status,
            errors=errors
        ))

    return ProductImportPreviewResponse(
        total_rows=len(rows_data),
        valid_rows=valid_count,
        invalid_rows=invalid_count,
        duplicate_rows=duplicate_count,
        new_categories=len(new_cats),
        rows=rows_data
    )

@router.post("/import/execute", response_model=ProductImportExecuteResponse)
async def execute_excel_import(
    request: ProductImportExecuteRequest,
    current_user: User = Depends(get_current_active_user),
    db: AsyncSession = Depends(get_db)
):
    valid_rows = [r for r in request.rows if r.status == "valid"]
    if not valid_rows:
        return ProductImportExecuteResponse(
            products_imported=0,
            categories_created=0,
            products_skipped=len(request.rows),
            products_failed=0
        )

    # 1. Resolve Categories
    unique_category_names = {r.category.strip() for r in valid_rows}
    
    existing_categories_result = await db.execute(
        select(Category).where(Category.owner_id == current_user.id)
    )
    existing_categories = existing_categories_result.scalars().all()
    
    # Create mapping by lowercase name
    cat_map = {c.name.lower(): c.id for c in existing_categories}
    
    categories_created = 0
    for cat_name in unique_category_names:
        lower_name = cat_name.lower()
        if lower_name not in cat_map:
            new_cat = Category(
                owner_id=current_user.id,
                name=cat_name,
                is_active=True
            )
            db.add(new_cat)
            await db.flush() # flush to get the ID
            cat_map[lower_name] = new_cat.id
            categories_created += 1

    # 2. Re-verify existing SKUs one last time
    existing_products_result = await db.execute(
        select(Product.sku).where(
            Product.owner_id == current_user.id,
            Product.sku.isnot(None)
        )
    )
    existing_skus = {sku.lower() for sku in existing_products_result.scalars().all()}
    
    # 3. Insert Products
    products_imported = 0
    products_skipped = len(request.rows) - len(valid_rows)
    products_failed = 0
    
    new_products = []
    
    for row in valid_rows:
        if row.sku and row.sku.lower() in existing_skus:
            products_skipped += 1
            continue
            
        try:
            product = Product(
                owner_id=current_user.id,
                name=row.name,
                sku=row.sku,
                category_id=cat_map[row.category.strip().lower()],
                purchase_price=row.cost_price,
                selling_price=row.selling_price,
                stock_quantity=row.stock,
                unit=row.unit,
                is_active=True
            )
            new_products.append(product)
            if row.sku:
                existing_skus.add(row.sku.lower())
            products_imported += 1
        except Exception:
            products_failed += 1

    if new_products:
        db.add_all(new_products)
        
    await db.commit()
    
    return ProductImportExecuteResponse(
        products_imported=products_imported,
        categories_created=categories_created,
        products_skipped=products_skipped,
        products_failed=products_failed
    )
